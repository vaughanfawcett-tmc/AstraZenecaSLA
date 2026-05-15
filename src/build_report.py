"""AstraZeneca SLA report builder.

Usage:
    python build_report.py <ticket_export.(xlsx|csv)> <YYYY-MM> [output.xlsx]

If output is omitted, writes to ../output/AstraZeneca_SLA_Report_<YYYY-MM>.xlsx.

Set OPENROUTER_API_KEY in the environment to enable Tier-2 LLM categorisation
for tickets the deterministic rules can't handle (~40% of rows).
"""
from __future__ import annotations
import sys, json, re, os
from datetime import datetime
from pathlib import Path
from collections import Counter, defaultdict

import calendar
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from categories import classify_tier1, CATEGORIES
from sla import sla_verdict
import tier2_classifier
import fresh_client


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent


COUNTRY_ALIASES = {
    # NB: Amy's report uses both "Czech" and "Czech Republic" inconsistently —
    # we leave Czech Republic alone rather than picking a winner.
    "United Kingdom": "UK",
    "Great Britain": "UK",
    "Deutschland": "Germany",
}

# Countries Amy includes in the report even when not in Jay's standard 25-country roster.
EXTRA_ALLOWED_COUNTRIES = {"Germany", "UK"}


def _normalise_country(c: str) -> str:
    c = (c or "").strip()
    return COUNTRY_ALIASES.get(c, c)


def load_email_country_map(extra_files: list[Path] | None = None) -> tuple[dict[str, str], set[str]]:
    """Load the email→country map and the allow-list of valid markets.

    Returns (email→country dict, allowed_countries set). Any country not in the
    allow-list (or unmapped emails) should be reported as "Other".

    The bundled `data/email_country.json` is optional — when not present (e.g.
    on a fresh hosted deploy where PII is not committed to the repo), the
    function returns an empty mapping and the caller is expected to provide
    `extra_files` with at least one country lookup. Without either source, every
    row falls into "Other".
    """
    mapping: dict[str, str] = {}
    allowed: set[str] = set()
    p = ROOT / "data" / "email_country.json"
    if p.exists():
        data = json.loads(p.read_text())
        mapping = {k: _normalise_country(v) for k, v in data["emails"].items()}
        allowed = {_normalise_country(c) for c in (data.get("allowed_countries") or [])}
    allowed |= EXTRA_ALLOWED_COUNTRIES
    if extra_files:
        for f in extra_files:
            extra = _load_extra_country_file(f)
            for k, v in extra.items():
                mapping[k.lower()] = _normalise_country(v)
                allowed.add(_normalise_country(v))
    return mapping, allowed


def _load_extra_country_file(path: Path) -> dict[str, str]:
    """Read an xlsx or csv file with email/country columns. Tolerant of column name variants."""
    out = {}
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        headers = [str(h or "").strip().lower() for h in rows[0]]
        email_idx = next((i for i, h in enumerate(headers) if "email" in h or "contact" in h), None)
        ctry_idx = next((i for i, h in enumerate(headers) if "country" in h or "market" in h), None)
        if email_idx is None or ctry_idx is None:
            return {}
        for r in rows[1:]:
            if email_idx < len(r) and ctry_idx < len(r):
                e = (r[email_idx] or "")
                c = (r[ctry_idx] or "")
                if e and c:
                    out[str(e).strip().lower()] = str(c).strip()
    elif path.suffix.lower() == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                e = c = None
                for k, v in row.items():
                    kl = (k or "").lower()
                    if not e and ("email" in kl or "contact" in kl):
                        e = v
                    if not c and ("country" in kl or "market" in kl):
                        c = v
                if e and c:
                    out[e.strip().lower()] = c.strip()
    return out


def _open_ticket_export(path: Path):
    """Return (headers, row_iterator) for an xlsx or csv ticket export.

    For CSV, "Created time" cells are parsed into datetime when they look like
    ISO-ish timestamps, so the month-filter and SLA logic see the same shape as
    the xlsx path (which openpyxl returns natively as datetime objects).
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        raw_iter = ws.iter_rows(values_only=True)
        headers = list(next(raw_iter))
        created_idx = next(
            (i for i, h in enumerate(headers) if (h or "").strip().lower() == "created time"),
            None,
        )

        def _xlsx_row_iter():
            for raw in raw_iter:
                if created_idx is not None and created_idx < len(raw) and not isinstance(raw[created_idx], datetime):
                    raw = list(raw)
                    raw[created_idx] = _parse_csv_datetime(raw[created_idx])
                    raw = tuple(raw)
                yield raw

        return headers, _xlsx_row_iter()

    if suffix == ".csv":
        import csv, io
        # Decode the whole file up-front with a fallback chain — Freshdesk exports
        # are usually utf-8 / utf-8-sig, but country-name columns sometimes carry
        # cp1252-encoded accents (e.g. "Türkiye"). Sniffing only the header row
        # would miss bad bytes deeper in the file.
        raw_bytes = path.read_bytes()
        text = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            # latin-1 can decode any byte sequence, so this branch is theoretically
            # unreachable — kept for defensive clarity.
            text = raw_bytes.decode("latin-1", errors="replace")

        reader = csv.reader(io.StringIO(text))
        try:
            headers = next(reader)
        except StopIteration:
            raise ValueError(f"CSV is empty: {path}")

        created_idx = next(
            (i for i, h in enumerate(headers) if (h or "").strip().lower() == "created time"),
            None,
        )

        def _row_iter():
            for raw in reader:
                if created_idx is not None and created_idx < len(raw):
                    raw[created_idx] = _parse_csv_datetime(raw[created_idx])
                yield tuple(_blank_to_none(v) for v in raw)

        return headers, _row_iter()

    raise ValueError(f"Unsupported ticket export format: {path.suffix}")


def _parse_csv_datetime(s):
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s
    text = str(s).strip()
    if not text:
        return None
    # Try common Freshdesk / spreadsheet timestamp formats.
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return text  # leave as-is; downstream isinstance() check will skip month filter


def _blank_to_none(v):
    if v is None:
        return None
    if isinstance(v, str) and v == "":
        return None
    return v


def is_astra_row(contact_id: str | None, full_name: str | None) -> bool:
    c = (contact_id or "").lower()
    f = (full_name or "").lower()
    return "astrazeneca" in c or "astrazeneca" in f


def parse_month(s: str) -> tuple[int, int]:
    y, m = s.split("-")
    return int(y), int(m)


def detect_month(ticket_export_path: Path) -> str | None:
    """Pick the most common (year, month) from Astra rows' Created time.

    Returns "YYYY-MM" or None if no usable datetimes were found.
    """
    headers, rows_iter = _open_ticket_export(ticket_export_path)
    H = {h: i for i, h in enumerate(headers) if h is not None}
    created_idx = H.get("Created time")
    contact_idx = H.get("Contact ID")
    name_idx = H.get("Full name")
    if created_idx is None:
        return None
    counts: Counter = Counter()
    for r in rows_iter:
        contact = r[contact_idx] if contact_idx is not None and contact_idx < len(r) else None
        full_name = r[name_idx] if name_idx is not None and name_idx < len(r) else None
        if not is_astra_row(contact, full_name):
            continue
        created = r[created_idx] if created_idx < len(r) else None
        if isinstance(created, datetime):
            counts[(created.year, created.month)] += 1
    if not counts:
        return None
    (y, m), _ = counts.most_common(1)[0]
    return f"{y:04d}-{m:02d}"


def build(
    ticket_export_path: Path,
    month: str,
    output_path: Path,
    extra_country_files: list[Path] | None = None,
    enable_tier2: bool | None = None,
    progress_cb=None,
) -> dict:
    """Build the SLA report.

    enable_tier2:
        None  → auto-detect from ANTHROPIC_API_KEY env var (default)
        True  → require it; raise if no key
        False → never call the LLM
    progress_cb: optional callback(stage: str, current: int, total: int, detail: str)
                 for UI progress updates.
    """
    year, mo = parse_month(month)
    email_country, allowed_countries = load_email_country_map(extra_country_files)

    if progress_cb: progress_cb("read", 0, 1, "loading ticket export")
    headers, rows_iter = _open_ticket_export(ticket_export_path)
    H = {h: i for i, h in enumerate(headers) if h is not None}

    required = {"Contact ID", "Full name", "Created time", "Ticket ID", "Subject"}
    missing = required - set(H)
    if missing:
        raise ValueError(
            f"This doesn't look like a Freshdesk ticket export — missing columns: "
            f"{', '.join(sorted(missing))}. "
            f"Make sure you upload the 'Ticket Export' file from Freshdesk, "
            f"not the country/roster lookup."
        )

    def col(row, name):
        idx = H.get(name)
        return row[idx] if idx is not None else None

    out_rows = []
    audit_rows = []
    tier2_pending = []  # indices in out_rows that need LLM classification
    unknown_emails = set()
    skipped_non_astra = 0
    skipped_wrong_month = 0
    now = datetime.now()

    for r in rows_iter:
        contact = col(r, "Contact ID")
        full_name = col(r, "Full name")
        if not is_astra_row(contact, full_name):
            skipped_non_astra += 1
            continue

        created = col(r, "Created time")
        if isinstance(created, datetime) and (created.year, created.month) != (year, mo):
            skipped_wrong_month += 1
            continue

        ticket_id = col(r, "Ticket ID")
        subject = col(r, "Subject") or ""
        qr = col(r, "Query Regarding")
        typ = col(r, "Type")
        source = col(r, "Source")
        full_name_clean = (full_name or "").strip()
        contact_clean = (contact or "").strip().lower()

        # Country — AstraZeneca uses a closed list of valid markets (loaded from
        # data/email_country.json). Unmapped emails, or values not on the allow-list,
        # are reported as "Other".
        country = email_country.get(contact_clean, "")
        if not country:
            fn_email = (full_name_clean or "").replace(" ", "").lower()
            country = email_country.get(fn_email, "")
        if not country:
            if contact_clean:
                unknown_emails.add(contact_clean)
            country = "Other"
        elif allowed_countries and country not in allowed_countries:
            country = "Other"

        # SLA verdict — tags lets us flip AutoClosed/customer-silent breaches to Within
        within, outside = sla_verdict(
            col(r, "Resolution status"),
            col(r, "First response status"),
            col(r, "Every response status"),
            created if isinstance(created, datetime) else None,
            now,
            tags=col(r, "Tags"),
        )

        # Tier 1 — when confidence < 0.90 we still capture the rule's answer as a
        # fallback but route the row to Tier-2 for LLM verification. Empirically
        # the low-confidence Tier-1 rules account for the bulk of Amy-disagreements.
        TIER1_TRUST_THRESHOLD = 0.90
        tier1 = classify_tier1(qr, typ, subject)
        if tier1 and tier1[1] >= TIER1_TRUST_THRESHOLD:
            (sub_cat, cat), confidence, reason = tier1
            tier = "tier1"
        elif tier1:
            (sub_cat, cat), confidence, reason = tier1
            tier = "tier2_needed"
            reason = f"low-conf tier1 ({confidence:.2f}); routing to Tier 2 — {reason}"
        else:
            sub_cat, cat, confidence, reason = "", "", 0.0, "no rule matched — needs Tier 2"
            tier = "tier2_needed"

        out_idx = len(out_rows)
        out_rows.append({
            "Month": datetime(year, mo, 1),
            "SUBCASE TITLE": subject,
            "Sub Category": sub_cat,
            "Category": cat,
            "CASE NUMBER": ticket_id,
            "CUSTOMER": full_name_clean,
            "EMAIL ADDRESS": contact_clean,
            "Country": country,
            "Within": within,
            "Outside": outside,
        })
        audit_rows.append({
            "case": ticket_id,
            "tier": tier,
            "confidence": confidence,
            "reason": reason,
            "qr": qr,
            "type": typ,
            "subject": subject[:60],
        })

        if tier == "tier2_needed":
            tier2_pending.append({
                "out_idx": out_idx,
                "ticket_id": ticket_id,
                "subject": subject,
                "source": source,
                "query_regarding": qr,
                "type": typ,
                "full_name": full_name_clean,
            })

    # Row-count sanity check — surface obvious problems before doing more work
    if not out_rows:
        raise ValueError(
            f"No AstraZeneca rows found for {month}. "
            f"Scanned {skipped_non_astra} non-Astra rows and skipped "
            f"{skipped_wrong_month} for wrong month. Check the export covers the "
            f"right month and contains AstraZeneca contacts."
        )

    # Tier 2 — LLM classification for the rows the rules couldn't handle
    use_tier2 = enable_tier2 if enable_tier2 is not None else tier2_classifier.is_available()
    tier2_classified = 0
    tier2_failed = 0
    fetch_stats: dict = {"fetched": 0, "empty": 0, "failed": 0}
    fresh = None
    if use_tier2 and tier2_pending and fresh_client.is_available():
        try:
            fresh = fresh_client.FreshClient()
        except Exception as e:
            print(f"[fresh] could not initialise client: {type(e).__name__}: {e}")
            fresh = None
    if use_tier2 and tier2_pending:
        if progress_cb:
            label = "calling Claude Haiku 4.5 (with Fresh transcripts)" if fresh else "calling Claude Haiku 4.5"
            progress_cb("tier2", 0, len(tier2_pending), label)

        def _on_progress(i, total, ticket_id):
            if progress_cb:
                progress_cb("tier2", i, total, f"ticket {ticket_id}")

        results = tier2_classifier.classify_batch(
            tier2_pending,
            on_progress=_on_progress,
            fresh=fresh,
            fetch_stats=fetch_stats,
        )
        for ticket, result in zip(tier2_pending, results):
            if result is None:
                tier2_failed += 1
                continue
            i = ticket["out_idx"]
            out_rows[i]["Sub Category"] = result.sub_category
            out_rows[i]["Category"] = result.category
            audit_rows[i]["tier"] = "tier2"
            audit_rows[i]["confidence"] = result.confidence
            audit_rows[i]["reason"] = f"LLM: {result.reasoning}"
            tier2_classified += 1

    # Build output workbook
    if progress_cb: progress_cb("write", 0, 1, "building xlsx")
    wbout = openpyxl.Workbook()
    ws_data = wbout.active
    ws_data.title = "Data"
    headers_out = ["Month", "SUBCASE TITLE", "Sub Category", "Category", "CASE NUMBER", "CUSTOMER", "EMAIL ADDRESS", "Country", "Within", "Outside"]
    ws_data.append(headers_out)
    bold = Font(bold=True)
    for c in ws_data[1]:
        c.font = bold

    review_fill = PatternFill("solid", fgColor="FFF4CC")  # yellow for low-confidence rows
    for i, row in enumerate(out_rows):
        ws_data.append([row[h] for h in headers_out])
        a = audit_rows[i]
        if a["tier"] == "tier2" and a["confidence"] < 0.6:
            for cell in ws_data[ws_data.max_row]:
                cell.fill = review_fill

    for r in range(2, ws_data.max_row + 1):
        ws_data.cell(row=r, column=1).number_format = "mmm-yy"

    # Categories sheet
    ws_cats = wbout.create_sheet("Categories")
    ws_cats.append(["Sub Category", "Category"])
    for c in ws_cats[1]: c.font = bold
    for s, c in CATEGORIES:
        ws_cats.append([s, c])

    # Quality Failures tab — matches Amy's shape: Month | Market | Quality Measure Failures | Quality Pass Rate
    ws_qf = wbout.create_sheet("Quality Failures")
    ws_qf.append(["Month", "Market", "Quality Measure Failures", "Quality Pass Rate"])
    for c in ws_qf[1]: c.font = bold

    by_country: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "outside": 0})
    for row in out_rows:
        ctry = row["Country"]
        by_country[ctry]["total"] += 1
        if row["Outside"] == "Y":
            by_country[ctry]["outside"] += 1

    month_name = calendar.month_name[mo]
    fail_fill = PatternFill("solid", fgColor="FFE6E6")
    # Always emit a row per allowed market (even with zero activity), then Other,
    # then any unexpected countries we saw. Matches Amy's per-month roster shape.
    markets = sorted(allowed_countries) + ["Other"]
    for ctry in markets + sorted(c for c in by_country if c not in markets):
        s = by_country.get(ctry, {"total": 0, "outside": 0})
        if s["total"] == 0 and ctry not in markets:
            continue
        pass_rate = ((s["total"] - s["outside"]) / s["total"]) if s["total"] else 1
        ws_qf.append([month_name, ctry, s["outside"], round(pass_rate, 4)])
        if s["outside"] > 0:
            for cell in ws_qf[ws_qf.max_row]:
                cell.fill = fail_fill
    # Format the pass rate column as a percent so Excel renders 0.95 → 95%
    for r in range(2, ws_qf.max_row + 1):
        ws_qf.cell(row=r, column=4).number_format = "0.00%"

    # Exceptions tab — tickets that need a human eye (low-conf Tier-2, Tier-2 failures,
    # rows that ended up with no category at all).
    ws_ex = wbout.create_sheet("Exceptions")
    ws_ex.append(["Case Number", "Reason", "Confidence", "Subject", "Query Regarding", "Type"])
    for c in ws_ex[1]: c.font = bold
    for i, a in enumerate(audit_rows):
        row = out_rows[i]
        cat = (row.get("Category") or "").strip()
        reason = None
        if not cat:
            reason = "No category assigned (Tier-2 failed or skipped)"
        elif a["tier"] == "tier2" and a["confidence"] < 0.6:
            reason = "Low-confidence Tier-2 classification"
        if reason:
            ws_ex.append([a["case"], reason, a["confidence"], row.get("SUBCASE TITLE", "")[:80], a["qr"], a["type"]])

    # Audit log
    ws_audit = wbout.create_sheet("Audit Log")
    ws_audit.append(["Case Number", "Tier", "Confidence", "Reason", "Query Regarding", "Type", "Subject"])
    for c in ws_audit[1]: c.font = bold
    for a in audit_rows:
        ws_audit.append([a["case"], a["tier"], a["confidence"], a["reason"], a["qr"], a["type"], a["subject"]])

    for ws in [ws_data, ws_cats, ws_qf, ws_ex, ws_audit]:
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wbout.save(output_path)

    # Persistent run log — append to output/runs.csv so we keep a history of every
    # report build. Header is written once on first run.
    try:
        runs_path = ROOT / "output" / "runs.csv"
        runs_path.parent.mkdir(parents=True, exist_ok=True)
        new_file = not runs_path.exists()
        with open(runs_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(["timestamp", "input_file", "month", "out_rows", "tier1", "tier2_ok",
                            "tier2_failed", "exceptions", "output_file", "status"])
            exceptions_count = ws_ex.max_row - 1  # minus header
            w.writerow([
                datetime.now().isoformat(timespec="seconds"),
                Path(ticket_export_path).name,
                month,
                len(out_rows),
                sum(1 for a in audit_rows if a["tier"] == "tier1"),
                tier2_classified,
                tier2_failed,
                exceptions_count,
                Path(output_path).name,
                "ok",
            ])
    except Exception as e:
        print(f"[runlog] could not append: {type(e).__name__}: {e}")

    return {
        "out_rows": len(out_rows),
        "skipped_non_astra": skipped_non_astra,
        "skipped_wrong_month": skipped_wrong_month,
        "unknown_emails": sorted(unknown_emails),
        "tier1_classified": sum(1 for a in audit_rows if a["tier"] == "tier1"),
        "tier2_classified": tier2_classified,
        "tier2_failed": tier2_failed,
        "tier2_needed_no_key": len(tier2_pending) if not use_tier2 else 0,
        "fresh_used": fresh is not None,
        "fresh_fetched": fetch_stats.get("fetched", 0),
        "fresh_empty": fetch_stats.get("empty", 0),
        "fresh_failed": fetch_stats.get("failed", 0),
    }


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    in_path = Path(sys.argv[1]).expanduser()
    month = sys.argv[2]
    out_path = Path(sys.argv[3]).expanduser() if len(sys.argv) >= 4 else ROOT / "output" / f"AstraZeneca_SLA_Report_{month}.xlsx"

    def cli_progress(stage, cur, total, detail):
        print(f"  [{stage}] {cur}/{total}  {detail}")

    stats = build(in_path, month, out_path, progress_cb=cli_progress)
    print(f"\n✓ Wrote {out_path}")
    print(f"  Astra rows in report : {stats['out_rows']}")
    print(f"  Tier-1 classified     : {stats['tier1_classified']}")
    print(f"  Tier-2 classified     : {stats['tier2_classified']}")
    print(f"  Tier-2 failed         : {stats['tier2_failed']}")
    print(f"  Need Tier 2 (no key)  : {stats['tier2_needed_no_key']}")
    if stats["unknown_emails"]:
        print(f"\n  ⚠ {len(stats['unknown_emails'])} email(s) without country mapping:")
        for e in stats["unknown_emails"][:20]:
            print(f"      {e}")


if __name__ == "__main__":
    main()
